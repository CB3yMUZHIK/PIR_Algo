import openpyxl
from openpyxl.chart import LineChart,Reference


wb = openpyxl.Workbook() # Создание книги exel
wb.create_sheet('Первый лист', index=0)
ws = wb.active  #Выбор активного листа


def FindStart(a, l, m0):
    for i in range(l):
        if a[i][1] != m0:
            return a[i][1]


# Массивы значений пористости, а также концентрации
Alpha = []
m = []
grads = []
# Константы
dt = 1
dx = 1
k = 10 ** (-13)
mu = 10 ** (-3)
PH0 = 1.5 * (10 ** 7)
Pg = 10 ** 7
l = 100
check = 0
t = 0
ast = 10 ^ (-7)
gamma1 = 0.2  # кольматация
gamma2 = 0.7  # суффозия
m0 = 0.5
a0 = 0.3
G = 1000
mst = 0.01
dPH = 100
# Формулы первоначальных значений
grad = (PH0 - Pg) / l
v = (k / mu) * grad
# Заполнение
for i in range(0, 100):
    Alpha.append([0, 0])
    m.append([0, 0])

for i in range(1, 100):
    Alpha[i][0] = 0
    m[i][0] = m0

m[0][0] = m0
Alpha[0][0] = a0
Alpha[0][1] = a0

# Вычисления
ws['A1'] = 'Пористость'
ws['B1'] = 'Концентрация нагнетательной жидкости'
number = 2 # номер строки в таблице
ws['A2'] = m0 #Ввод значения в эксель

while t <= 100000:
    PH = PH0 + dPH * t
    grad = (PH - Pg) / l
    grads.append(grad)
    v = (k / mu) * grad
    if m0 <= mst:
        m[0][1] = mst
    else:
        m[0][1] = (-gamma1 * Alpha[0][0] * (m[0][0] - mst)) * dt + m[0][0]
        if grad > G:
            m[0][1] += gamma2 * (m0 - m[0][0]) * (grad - G)

    for i in range(1, 100):
        if m[i][0] <= mst:
            m[i][1] = mst
            Alpha[i][1] = (-v * (Alpha[i][0] - Alpha[i - 1][0]) / dx) * (dt / m[i][0]) + Alpha[i][0]
        else:
            m[i][1] = (-gamma1 * Alpha[i][0] * (m[i][0] - mst)) * dt + m[i][0]
            if grad > G:
                m[i][1] += gamma2 * (m0 - m[i][0]) * (grad)
            Alpha[i][1] = (-v * (Alpha[i][0] - Alpha[i - 1][0]) / dx) *(dt / m[i][0]) + Alpha[i][0]
        if Alpha[i][1] <= ast:
            Alpha[i][1] = ast
    for i in range(0, 100):
        Alpha[i][0] = Alpha[i][1]
        m[i][0] = m[i][1]
        check = m[i][0]
    t = t + dt

for i in range(2,l+2):
    secondnameA = 'A' + str(i)
    secondnameB = 'B' + str(i)
    ws[secondnameA] = m[i-2][0]
    ws[secondnameB] = Alpha[i-2][0]


wb.save('test.xlsx') # сохранение файла
# cоздание диаграммы
sheet = wb['Первый лист']
# min max столбцов и строк таблицы
min_column = 1
max_column = 1
min_row = wb.active.min_row
max_row = wb.active.max_row
linechart = LineChart()
linechart.title = 'Пористость'
data = Reference(sheet, min_col = min_column, min_row = min_row, max_col = max_column, max_row = max_row)
linechart.add_data(data, titles_from_data=True)
sheet.add_chart(linechart, "K12")
min_column = 2
max_column = 2
min_row = wb.active.min_row
max_row = wb.active.max_row
linechart = LineChart()
linechart.title = 'Концентрация нагнетательной жидкости'
data = Reference(sheet, min_col = min_column, min_row = min_row, max_col = max_column, max_row = max_row)
linechart.add_data(data, titles_from_data=True)
sheet.add_chart(linechart, "K30")
wb.save('test.xlsx')


