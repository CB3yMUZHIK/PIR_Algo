import openpyxl
from tkinter import*
from tkinter import messagebox
from openpyxl.chart import LineChart,Reference

def main(dt,dx,k,mu,PH0,Pg,l,check,t,ast,gamma1,gamma2,m0,a0,G,mst,dPH):
    # Массивы значений пористости, а также концентрации
    Alpha = []
    m = []
    grads = []

    dt = dt
    dx = dx
    k = k
    mu = mu
    PH0 = PH0
    Pg = Pg
    l = l
    check = check
    t = t
    ast = ast
    gamma1 = gamma1  # кольматация
    gamma2 = gamma2  # суффозия
    m0 = m0
    a0 = a0
    G = G
    mst = mst
    dPH = dPH
    
    grad = (PH0 - Pg) / l
    v = (k / mu) * grad
    # Заполнение
    for i in range(0, 100):
        Alpha.append([0, 0])
        m.append([0, 0])

    for i in range(1, 100):
        Alpha[i][0] = 0
        m[i][0] = m0

    m[0][0] = m0
    Alpha[0][0] = a0
    Alpha[0][1] = a0

    # Вычисления
    ws['A1'] = 'Пористость'
    ws['B1'] = 'Концентрация частиц примеси'
    number = 2 # номер строки в таблице
    ws['A2'] = m0 #Ввод значения в эксель

    while t <= 100000:
        PH = PH0 + dPH * t
        grad = (PH - Pg) / l
        grads.append(grad)
        v = (k / mu) * grad
        if m0 <= mst:
            m[0][1] = mst
        else:
            m[0][1] = (-gamma1 * Alpha[0][0] * (m[0][0] - mst)) * dt + m[0][0]
            if grad > G:
                m[0][1] += gamma2 * (m0 - m[0][0]) * (grad - G)

        for i in range(1, 100):
            if m[i][0] <= mst:
                m[i][1] = mst
                Alpha[i][1] = (-v * (Alpha[i][0] - Alpha[i - 1][0]) / dx) * (dt / m[i][0]) + Alpha[i][0]
            else:
                m[i][1] = (-gamma1 * Alpha[i][0] * (m[i][0] - mst)) * dt + m[i][0]
                if grad > G:
                    m[i][1] += gamma2 * (m0 - m[i][0]) * (grad)
                Alpha[i][1] = (-v * (Alpha[i][0] - Alpha[i - 1][0]) / dx) *(dt / m[i][0]) + Alpha[i][0]
            if Alpha[i][1] <= ast:
                Alpha[i][1] = ast
        for i in range(0, 100):
            Alpha[i][0] = Alpha[i][1]
            m[i][0] = m[i][1]
            check = m[i][0]
        t = t + dt

    for i in range(2,l+2):
        secondnameA = 'A' + str(i)
        secondnameB = 'B' + str(i)
        ws[secondnameA] = m[i-2][0]
        ws[secondnameB] = Alpha[i-2][0]


    wb.save('test.xlsx') # сохранение файла
    # cоздание диаграммы
    sheet = wb['Первый лист']
    # min max столбцов и строк таблицы
    min_column = 1
    max_column = 1
    min_row = wb.active.min_row
    max_row = wb.active.max_row
    linechart = LineChart()
    linechart.title = 'Пористость'
    data = Reference(sheet, min_col = min_column, min_row = min_row, max_col = max_column, max_row = max_row)
    linechart.add_data(data, titles_from_data=True)
    sheet.add_chart(linechart, "K12")
    min_column = 2
    max_column = 2
    min_row = wb.active.min_row
    max_row = wb.active.max_row
    linechart = LineChart()
    linechart.title = 'Концентрация нагнетательной жидкости'
    data = Reference(sheet, min_col = min_column, min_row = min_row, max_col = max_column, max_row = max_row)
    linechart.add_data(data, titles_from_data=True)
    sheet.add_chart(linechart, "K30")
    wb.save('test.xlsx')

    messagebox.showinfo('ВСЕ','ВСЕ')


def clicked():
    # Константы
    dt = 1
    dx = 1
    k = 10 ** (-13)
    mu = 10 ** (-3)
    PH0 = 1.5 * (10 ** 7)
    Pg = 10 ** 7
    l = 100
    check = 0
    t = 0
    ast = 10 ^ (-7)
    gamma1 = 0.2  # кольматация
    gamma2 = 0.7  # суффозия
    m0 = 0.5
    a0 = 0.3
    G = 1000
    mst = 0.01
    dPH = 100
    main(dt,dx,k,mu,PH0,Pg,l,check,t,ast,gamma1,gamma2,m0,a0,G,mst,dPH)


wb = openpyxl.Workbook() # Создание книги exel
wb.create_sheet('Первый лист', index=0)
ws = wb.active  #Выбор активного листа


window = Tk()
window.geometry('500x500')
window.title("НАШ ПРОЕКТ")
lbl = Label(window, text = "Создание графика с Вашими значениями")
btn = Button(window, text = "Создать график с дефолтными значениями", command = clicked)
#Текстовые поля для задачи переменных
txt_dt = Entry(window, width = 10)
txt_dx = Entry(window, width = 10)
txt_k = Entry(window, width = 10)
txt_mu = Entry(window, width = 10)
txt_PH0 = Entry(window, width = 10)
txt_Pg = Entry(window, width = 10)
txt_l = Entry(window, width = 10)
txt_check = Entry(window, width = 10)
txt_t = Entry(window, width = 10)
txt_ast = Entry(window, width = 10)
txt_gamma1 = Entry(window, width = 10)
txt_gamma2 = Entry(window, width = 10)
txt_m0 = Entry(window, width = 10)
txt_a0 = Entry(window, width = 10)
txt_G = Entry(window, width = 10)
txt_mst = Entry(window, width = 10)
txt_dPH = Entry(window, width = 10)
#Подписи к полям
lbl_dt = Label(window, text = "dt")
lbl_dx = Label(window, text = "dx")
lbl_k = Label(window, text = "k")
lbl_mu = Label(window, text = "mu")
lbl_PH0 = Label(window, text = "PH0")
lbl_Pg = Label(window, text = "Pg")
lbl_l = Label(window, text = "l")
lbl_check = Label(window, text = "check")
lbl_t = Label(window, text = "t")
lbl_ast = Label(window, text = "ast")
lbl_gamma1 = Label(window, text = "gamma1")
lbl_gamma2 = Label(window, text = "gamma2")
lbl_m0 = Label(window, text = "m0")
lbl_a0 = Label(window, text = "a0")
lbl_G = Label(window, text = "G")
lbl_mst = Label(window, text = "mst")
lbl_dPH = Label(window, text = "dPH")
#Расположение полей в окне
btn.grid(column = 0, row = 0)
lbl.grid(column = 0, row = 1)
txt_dt.grid(column = 1, row = 2)
txt_dx.grid(column = 1, row = 3)
txt_k.grid(column = 1, row = 4)
txt_mu.grid(column = 1, row = 5)
txt_PH0.grid(column = 1, row = 6)
txt_Pg.grid(column = 1, row = 7)
txt_l.grid(column = 1, row = 8)
txt_check.grid(column = 1, row = 9)
txt_t.grid(column = 1, row = 10)
txt_ast.grid(column = 1, row = 11)
txt_gamma1.grid(column = 1, row = 12)
txt_gamma2.grid(column = 1, row = 13)
txt_m0.grid(column = 1, row = 14)
txt_a0.grid(column = 1, row = 15)
txt_G.grid(column = 1, row = 16)
txt_mst.grid(column = 1, row = 17)
txt_dPH.grid(column = 1, row = 18)
lbl_dt.grid(column = 0, row = 2)
lbl_dx.grid(column = 0, row = 3)
lbl_k.grid(column = 0, row = 4)
lbl_mu.grid(column = 0, row = 5)
lbl_PH0.grid(column = 0, row = 6)
lbl_Pg.grid(column = 0, row = 7)
lbl_l.grid(column = 0, row = 8)
lbl_check.grid(column = 0, row = 9)
lbl_t.grid(column = 0, row = 10)
lbl_ast.grid(column = 0, row = 11)
lbl_gamma1.grid(column = 0, row = 12)
lbl_gamma2.grid(column = 0, row = 13)
lbl_m0.grid(column = 0, row = 14)
lbl_a0.grid(column = 0, row = 15)
lbl_G.grid(column = 0, row = 16)
lbl_mst.grid(column = 0, row = 17)
lbl_dPH.grid(column = 0, row = 18)
window.mainloop()


def FindStart(a, l, m0):
    for i in range(l):
        if a[i][1] != m0:
            return a[i][1]


