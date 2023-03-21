import openpyxl
import os.path
wb = openpyxl.Workbook()
ws = wb.active

def FindStart(a, l, m0):
    for i in range(l):
        if a[i][1] != m0:
            return a[i][1]


# Массивы значений пористости, а также концентрации
Alpha = []
m = []
# Константы
dt = 0.1
dx = 1
k = 10 ** (-13)
mu = 10 ** (-3)
PH0 = 1.5 * (10 ** 7)
Pg = 10 ** 7
l = 100
check = 0
t = 0
ast = 10 ^ (-7)
gamma1 = 0.5  # кольматация
gamma2 = 0.01  # суффозия
m0 = 0.5
a0 = 0.3
G = 50200
mst = 0.01
dPH = 100
# Формулы первоначальных значений
grad = (PH0 - Pg) / l
v = (k / mu) * grad
# Заполнение
for i in range(0, 100):
    Alpha.append([0, 0])
    m.append([0, 0])

for i in range(1, 100):
    Alpha[i][0] = 0
    m[i][0] = m0

m[0][0] = m0
Alpha[0][0] = a0
Alpha[0][1] = a0

# Вычисления
number = 2 # номер строки в таблице
ws['A2'] = m0 #Ввод значания в эксель
while t <= 100000:
    PH = PH0 + dPH * t
    grad = (PH - Pg) / l
    v = (k / mu) * grad
    if m0 <= mst:
        m[0][1] = mst
    else:
        m[0][1] = (-gamma1 * Alpha[0][0] * (m[0][0] - mst)) * dt + m[0][0]
        if grad > G:
            m[0][1] += gamma2 * (m0 - (m[0][0] - mst) * (grad - G))

    for i in range(1, 100):
        if m[i][0] <= mst:
            m[i][1] = mst
            Alpha[i][1] = (-v * (Alpha[i][0] - Alpha[i - 1][0]) / dx) * (dt / m[i][0]) + Alpha[i][0]
        else:
            m[i][1] = (-gamma1 * Alpha[i][0] * (m[i][0] - mst)) * dt + m[i][0]
            if grad > G:
                m[i][1] += gamma2 * (m0 - (m[i][0] - mst) * (grad - G))
            Alpha[i][1] = Alpha[i][0] + (dt / m[i][0]) * (
                        (-(1 - Alpha[i][0]) * gamma1 * Alpha[i][0] * (m[i][0] - mst)) - v * (
                            Alpha[i][0] - Alpha[i - 1][0]) / dx)
    for i in range(0, 100):
        Alpha[i][0] = Alpha[i][1]
        m[i][0] = m[i][1]
        check = m[i][0]
    if m[0][1] > 1:
        break
    elif m[0][1] >= 0:
        number += 1
        print(number)
        secondnameA = 'A' + str(number)
        secondnameB = 'B' + str(number)
        ws[secondnameA] = m[0][1]
        ws[secondnameB] = grad
    t = t + dt

wb.save('test.xlsx')
