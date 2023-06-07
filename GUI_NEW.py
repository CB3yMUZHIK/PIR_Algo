import openpyxl
from tkinter import*
from tkinter import messagebox
from openpyxl.chart import LineChart, Reference


def FindStart(a, l, m0):
    for i in range(l):
        if a[i][1] != m0:
            return a[i][1]


def main(dt, dx, k, mu, PH0, Pg, l, check, t, ast, gamma1, gamma2, m0, a0, G, mst, dPH): # Основная функция для расчетов и создания файла
    wb = openpyxl.Workbook()  # Создание книги exel
    wb.create_sheet('Первый лист', index=0)
    ws = wb.active  # Выбор активного листа

    # Массивы значений пористости, а также концентрации
    Alpha = []
    m = []
    grads = []

    dt = dt
    dx = dx
    k = k
    mu = mu
    PH0 = PH0
    Pg = Pg
    l = l
    check = check
    t = t
    ast = ast
    gamma1 = gamma1  # кольматация
    gamma2 = gamma2  # суффозия
    m0 = m0
    a0 = a0
    G = G
    mst = mst
    dPH = mu

    grad = (PH0 - Pg) / l
    v = (k / mu) * grad
    # Заполнение
    for i in range(0, 100):
        Alpha.append([0, 0])
        m.append([0, 0])

    for i in range(1, 100):
        Alpha[i][0] = 0
        m[i][0] = m0

    m[0][0] = m0
    Alpha[0][0] = a0
    Alpha[0][1] = a0

    # Вычисления
    ws['A1'] = 'Пористость'
    ws['B1'] = 'Концентрация нагнетательной жидкости'
    number = 2  # номер строки в таблице
    ws['A2'] = m0  # Ввод значения в эксель

    while t <= 100000:
        PH = PH0 + dPH * t
        grad = (PH - Pg) / l
        grads.append(grad)
        v = (k / mu) * grad
        if m0 <= mst:
            m[0][1] = mst
        else:
            m[0][1] = (-gamma1 * Alpha[0][0] * (m[0][0] - mst)) * dt + m[0][0]
            if grad > G:
                m[0][1] += gamma2 * (m0 - m[0][0]) * (grad - G)

        for i in range(1, 100):
            if m[i][0] <= mst:
                m[i][1] = mst
                Alpha[i][1] = (-v * (Alpha[i][0] - Alpha[i - 1][0]) / dx) * (dt / m[i][0]) + Alpha[i][0]
            else:
                m[i][1] = (-gamma1 * Alpha[i][0] * (m[i][0] - mst)) * dt + m[i][0]
                if grad > G:
                    m[i][1] += gamma2 * (m0 - m[i][0]) * (grad)
                Alpha[i][1] = (-v * (Alpha[i][0] - Alpha[i - 1][0]) / dx) * (dt / m[i][0]) + Alpha[i][0]
            if Alpha[i][1] <= ast:
                Alpha[i][1] = ast
        for i in range(0, 100):
            Alpha[i][0] = Alpha[i][1]
            m[i][0] = m[i][1]
            check = m[i][0]
        t = t + dt

    for i in range(2, l + 2):
        secondnameA = 'A' + str(i)
        secondnameB = 'B' + str(i)
        ws[secondnameA] = m[i - 2][0]
        ws[secondnameB] = Alpha[i - 2][0]

    wb.save('test.xlsx')  # сохранение файла
    # cоздание диаграммы
    sheet = wb['Первый лист']
    # min max столбцов и строк таблицы
    min_column = 1
    max_column = 1
    min_row = wb.active.min_row
    max_row = wb.active.max_row
    linechart1 = LineChart()
    linechart1.y_axis.title = "Y"
    linechart1.x_axis.title = "X"
    linechart1.title = 'Пористость'
    data = Reference(sheet, min_col=min_column, min_row=min_row, max_col=max_column, max_row=max_row)
    linechart1.add_data(data, titles_from_data=True)
    sheet.add_chart(linechart1, "K12")
    min_column = 2
    max_column = 2
    min_row = wb.active.min_row
    max_row = wb.active.max_row
    linechart2 = LineChart()
    linechart2.y_axis.title = "Y"
    linechart2.x_axis.title = "X"
    linechart2.title = 'Концентрация нагнетательной жидкости'
    data = Reference(sheet, min_col=min_column, min_row=min_row, max_col=max_column, max_row=max_row)
    linechart2.add_data(data, titles_from_data=True)
    sheet.add_chart(linechart2, "K30")
    wb.save('test.xlsx')

    messagebox.showinfo('Уведомление', 'Проверьте Ваш рабочий стол') # Непонятно где сохраняется файл


def defolt(): # Функция с заданными параметрами
    # Константы
    dt = 1
    dx = 1
    k = 10 ** (-13)
    mu = 10 ** (-3)
    PH0 = 1.5 * (10 ** 7)
    Pg = 10 ** 7
    l = 100
    check = 0
    t = 0
    ast = 10 ^ (-7)
    gamma1 = 0.2  # кольматация
    gamma2 = 0.7  # суффозия
    m0 = 0.5
    a0 = 0.3
    G = 1000
    mst = 0.01
    dPH = mu
    main(dt, dx, k, mu, PH0, Pg, l, check, t, ast, gamma1, gamma2, m0, a0, G, mst, dPH)


def uservalues(): # Функция с параметрами пользователя
    dt = txt_dt.get()
    dx = txt_dx.get()
    k = txt_k.get()
    mu = txt_mu.get()
    PH0 = txt_PH0.get()
    Pg = txt_Pg.get()
    l = txt_l.get()
    check = txt_check.get()
    t = txt_t.get()
    ast = txt_ast.get()
    gamma1 = txt_gamma1.get()
    gamma2 = txt_gamma2.get()
    m0 = txt_m0.get()
    a0 = txt_a0.get()
    G = txt_G.get()
    mst = txt_mst.get()
    dPH = mu
    # Проверка на пустые значения
    if(dt == '' or dx == '' or k == '' or mu == '' or PH0 == '' or Pg == '' or l == '' or check == '' or t == '' or ast == '' or gamma1 == '' or gamma2 == '' or m0 == '' or a0 == ''  or G == '' or mst == '' or dPH == ''):
        messagebox.showinfo('Уведомление', 'Вы ввели не все значения')
    else:
        dt = float(txt_dt.get())
        dx = float(txt_dx.get())
        k = float(txt_k.get())
        mu = float(txt_mu.get())
        PH0 = float(txt_PH0.get())
        Pg = float(txt_Pg.get())
        l = int(txt_l.get())  # Цикл не работает, если сделать float
        check = float(txt_check.get())
        t = float(txt_t.get())
        ast = float(txt_ast.get())
        gamma1 = float(txt_gamma1.get())
        gamma2 = float(txt_gamma2.get())
        m0 = float(txt_m0.get())
        a0 = float(txt_a0.get())
        G = float(txt_G.get())
        mst = float(txt_mst.get())
        dPH = float(mu)
        main(dt,dx,k,mu,PH0,Pg,l,check,t,ast,gamma1,gamma2,m0,a0,G,mst,dPH)


# Интерфейс
window = Tk()
window.geometry('850x700')
window.title("НАШ ПРОЕКТ") #Надо поменять, наверное
lbl = Label(window, text="Создание графика с Вашими значениями:", font=("Arial Bold", 17))
wrg = Label(window, text="Примечание: для разделения целой части и дробной в десятичных дробях используйте «.» - (точку)", fg="red", font=10)
btn = Button(window, text="Создать график с предустановленными значениями", command=defolt, font=("Arial", 14))
btn_second = Button(window, text="Создать график", command=uservalues, font=("Arial", 10))
# Текстовые поля для задачи переменных
txt_dt = Entry(window, width=15)
txt_k = Entry(window, width=15)
txt_mu = Entry(window, width=15)
txt_PH0 = Entry(window, width=15)
txt_Pg = Entry(window, width=15)
txt_l = Entry(window, width=15)
txt_t = Entry(window, width=15)
txt_ast = Entry(window, width=15)
txt_gamma1 = Entry(window, width=15)
txt_gamma2 = Entry(window, width=15)
txt_m0 = Entry(window, width=15)
txt_a0 = Entry(window, width=15)
txt_G = Entry(window, width=15)
txt_mst = Entry(window, width=15)
# Подписи к полям
lbl_dt = Label(window, text="Шаг времени (с)", font=("Arial", 15))
lbl_k = Label(window, text="Площадь пласта (м^2)", font=("Arial", 15))
lbl_mu = Label(window, text="Наращивание нагнетательного давления (Па)", font=("Arial", 15))
lbl_PH0 = Label(window, text="Начальное нагнетательное давление (Па)", font=("Arial", 15))
lbl_Pg = Label(window, text="Начальное давление пласта (Па)", font=("Arial", 15))
lbl_l = Label(window, text="Длина пласта (м)", font=("Arial", 15))
lbl_t = Label(window, text="Время, данное процессу (с)", font=("Arial", 15))
lbl_ast = Label(window, text="Минимальное значение концентрации жидкости", font=("Arial", 15))
lbl_gamma1 = Label(window, text="Коэффициент кольматации", font=("Arial", 15))
lbl_gamma2 = Label(window, text="Коэффициент суффозии", font=("Arial", 15))
lbl_m0 = Label(window, text="Начальная пористость", font=("Arial", 15))
lbl_a0 = Label(window, text="Начальная концентрация нагнетательной жидкости", font=("Arial", 15))
lbl_G = Label(window, text="Предельное значение депрессии", font=("Arial", 15))
lbl_mst = Label(window, text="Минимальное значение пористости", font=("Arial", 15))
# Расположение полей в окне
wrg.grid(column=0,row=20)
btn.grid(column=0, row=0)
btn_second.grid(column=1, row=19)
lbl.grid(column=0, row=1)
txt_dt.grid(column=1, row=2)
txt_k.grid(column=1, row=4)
txt_mu.grid(column=1, row=5)
txt_PH0.grid(column=1, row=6)
txt_Pg.grid(column=1, row=7)
txt_l.grid(column=1, row=8)
txt_t.grid(column=1, row=10)
txt_ast.grid(column=1, row=11)
txt_gamma1.grid(column=1, row=12)
txt_gamma2.grid(column=1, row=13)
txt_m0.grid(column=1, row=14)
txt_a0.grid(column=1, row=15)
txt_G.grid(column=1, row=16)
txt_mst.grid(column=1, row=17)
lbl_dt.grid(column=0, row=2)
lbl_k.grid(column=0, row=4)
lbl_mu.grid(column=0, row=5)
lbl_PH0.grid(column=0, row=6)
lbl_Pg.grid(column=0, row=7)
lbl_l.grid(column=0, row=8)
lbl_t.grid(column=0, row=10)
lbl_ast.grid(column=0, row=11)
lbl_gamma1.grid(column=0, row=12)
lbl_gamma2.grid(column=0, row=13)
lbl_m0.grid(column=0, row=14)
lbl_a0.grid(column=0, row=15)
lbl_G.grid(column=0, row=16)
lbl_mst.grid(column=0, row=17)
window.mainloop()


